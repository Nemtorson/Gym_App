from turtle import goto
import openpyxl
from datetime import datetime 
from genericpath import exists
import tkinter as tk
from tkinter import filedialog, Text, Entry, Button, Label
from PIL import Image, ImageTk
import os
import re


global path
path = r"<insert path to where your excel file would be"


def create_base():
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        print("File with data already exists")
    
    except:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
         
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Name'
        sheet['C1'] = 'Series'
        sheet['D1'] = 'Reps'
        sheet['E1'] = 'Weight'
        print("New file created")
    workbook.save('gym_app_data.xlsx')
#if its possible to open an existing file, then dont create a new one
create_base()


def read_value(date):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    i = 2
    for cell in sheet.iter_rows(min_row = 1, min_col = 1, max_row = sheet.max_row, max_col = 1):
        cell = sheet.cell(i, 1).value
        if cell == date:
            for row in sheet.iter_rows(min_row = i, min_col = 1, max_row = sheet.max_row, max_col = 5):
                print(f"Name: {row[0].offset(column=1).value}, "
                          f"Series: {row[0].offset(column=2).value}, Reps: {row[0].offset(column=3).value}, "
                          f"Weight: {row[0].offset(column=4).value}")
                break
        i += 1        
#format dd.mm.yyyy   
    

#read_value("09.12.2023")


def add_value(name, series, reps, weights):

    global input
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    row_ins = sheet.max_row + 1
    data = [datetime.now().strftime("%d.%m.%Y"), name, series, reps, weights]
    input = data
    print(data[1],data[2],data[3],data[4])

    if data[0] != sheet.cell(row = sheet.max_row, column = 1).value:
        sheet.cell(row = row_ins, column = 1, value = "Poczatek Treningu")
        sheet.cell(row = row_ins + 1, column = 1, value = data[0])
        sheet.cell(row = row_ins + 1, column = 2, value = data[1])
        sheet.cell(row = row_ins + 1, column = 3, value = data[2])
        sheet.cell(row = row_ins + 1, column = 4, value = data[3])
        sheet.cell(row = row_ins + 1, column = 5, value = data[4])
    
    else:
            sheet.cell(row = row_ins, column = 1, value = data[0])
            sheet.cell(row = row_ins, column = 2, value = data[1])
            sheet.cell(row = row_ins, column = 3, value = data[2])
            sheet.cell(row = row_ins, column = 4, value = data[3])
            sheet.cell(row = row_ins, column = 5, value = data[4])

    workbook.save(path)
    message = "LIGHT WEIGHT BABYYYYY"
    display_label.config(text=message)


def add_value_help():
    return lambda : add_value(str(entry_name.get()), float(entry_series.get()), float(entry_reps.get()), float(entry_weights.get()))

def read_value_help():
    return lambda : read_value(str(entry_date.get()))

def center_image(canvas, img):
    canvas_width = canvas.winfo_width()
    canvas_height = canvas.winfo_height()

    x = (canvas_width - img.width()) / 2
    y = (canvas_height - img.height()) / 2

    canvas.create_image(x, y, anchor=tk.NW, image=img)

def on_canvas_resize(event):
    canvas.delete("all")
    center_image(canvas, background_image)

root = tk.Tk()
root.title("Gym_App")

frame = tk.Frame(root)
frame.pack()

image = Image.open("ronnie_coleman.jpg")  # Replace with the actual path to your image
background_image = ImageTk.PhotoImage(image)


label_name = Label(frame, text = "Name:")
label_name.pack()

entry_name = Entry(frame)
entry_name.pack()

label_series = Label(frame, text = "Series:")
label_series.pack()

entry_series = Entry(frame)
entry_series.pack()

label_reps = Label(frame, text = "Reps:")
label_reps.pack()

entry_reps = Entry(frame)
entry_reps.pack()

label_weights = Label(frame, text = "Weights:")
label_weights.pack()

entry_weights = Entry(frame)
entry_weights.pack()

button_add = Button(frame, text = "INSERT THEM GAINS BOI", command = add_value_help())
button_add.pack()

entry_date = Entry(frame)
entry_date.pack()

button_read = Button(frame, text = "See training on date", command = read_value_help())
button_read.pack()

display_label = Label(frame, text = "", font = ("Arial", 12))
display_label.pack()

canvas = tk.Canvas(frame, height = 600, width = 1200, bg="#131314")
canvas.pack(expand = True, fill = tk.BOTH)

canvas.bind("<Configure>", on_canvas_resize)


root.mainloop()
